from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ✅ Read student names from selected Excel file
def get_students(file_name):
    if not os.path.exists(file_name):
        return []
    wb = load_workbook(file_name)
    ws = wb.active
    names = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            names.append(name)
    wb.close()
    return names

# ✅ Mark attendance (P/A) for the selected Excel file
def mark_attendance(file_name, present_students):
    today = datetime.now().strftime('%Y-%m-%d')

    if not os.path.exists(file_name):
        return "File not found. Please add students first."

    wb = load_workbook(file_name)
    ws = wb.active

    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == today:
            wb.close()
            return "Attendance already marked"

    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = today

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            ws.cell(row=row, column=new_col).value = "P" if name in present_students else "A"

    wb.save(file_name)
    wb.close()
    return "Marked successfully"

# ✅ Merge all student names into a single Excel file
def merge_all_students(output_file="all_students.xlsx"):
    class_files = ["ClassA.xlsx", "ClassB.xlsx", "ClassC.xlsx"]
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "All Students"
    ws_out.append(["Name", "Class File"])

    for file in class_files:
        if not os.path.exists(file):
            continue
        wb = load_workbook(file)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if name:
                ws_out.append([name, file])
        wb.close()

    wb_out.save(output_file)
    return output_file