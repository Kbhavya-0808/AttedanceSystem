# attendance_app.py (updated with dynamic class creation and dropdown refresh)

import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
import glob
from attendance_utils import get_students, mark_attendance, merge_all_students

# Get all .xlsx class files except merged one
def get_class_files():
    return [f for f in glob.glob("*.xlsx") if f != "all_students.xlsx"]

def refresh_dropdown():
    class_dropdown['values'] = get_class_files()
    if get_class_files():
        class_var.set(get_class_files()[0])

# Add a student to the selected file
def add_student_to_excel(name, file_name):
    if not name:
        messagebox.showwarning("Input Error", "Please enter a student name.")
        return

    if not file_name:
        messagebox.showwarning("Input Error", "Please select a class file.")
        return

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.cell(row=1, column=1).value = "Name"
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    ws.cell(row=row, column=1).value = name
    wb.save(file_name)
    wb.close()
    messagebox.showinfo("Success", f"{name} added to {file_name}")

# GUI Setup
root = tk.Tk()
root.title("Attendance System")
root.geometry("500x700")

# Create new class file
tk.Label(root, text="Create New Class File:", font=("Arial", 12)).pack(pady=10)
new_class_entry = tk.Entry(root, font=("Arial", 12), width=30)
new_class_entry.pack()

def create_class_file():
    name = new_class_entry.get().strip()
    if not name.endswith(".xlsx"):
        name += ".xlsx"

    if os.path.exists(name):
        messagebox.showinfo("Exists", f"{name} already exists.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.cell(row=1, column=1).value = "Name"
        wb.save(name)
        messagebox.showinfo("Created", f"{name} created successfully.")

    new_class_entry.delete(0, tk.END)
    refresh_dropdown()

tk.Button(root, text="Create Class", command=create_class_file, bg="darkblue", fg="white", font=("Arial", 12)).pack(pady=5)

# File selection
tk.Label(root, text="Select Class File:", font=("Arial", 12)).pack(pady=10)
class_var = tk.StringVar()
class_dropdown = ttk.Combobox(root, textvariable=class_var, values=get_class_files(), state="readonly", width=30)
class_dropdown.pack()
refresh_dropdown()

# Student Name Entry for Adding
tk.Label(root, text="Enter New Student Name:", font=("Arial", 12)).pack(pady=10)
add_entry = tk.Entry(root, font=("Arial", 12), width=30)
add_entry.pack()

def add_name():
    name = add_entry.get().strip()
    file = class_var.get().strip()
    add_student_to_excel(name, file)
    add_entry.delete(0, tk.END)

# Load and mark attendance section
checkbox_frame = tk.Frame(root)
checkbox_frame.pack(pady=10)
checkboxes = {}

def load_students():
    for widget in checkbox_frame.winfo_children():
        widget.destroy()
    checkboxes.clear()

    selected_file = class_var.get()
    if not selected_file:
        messagebox.showwarning("Error", "Please select a file.")
        return

    students = get_students(selected_file)
    if not students:
        messagebox.showwarning("No Data", "No students found in selected file.")
        return

    tk.Label(checkbox_frame, text="Select Present Students:", font=("Arial", 12)).pack()
    for student in students:
        var = tk.IntVar()
        cb = tk.Checkbutton(checkbox_frame, text=student, variable=var, font=("Arial", 10))
        cb.pack(anchor="w")
        checkboxes[student] = var

def submit_attendance():
    selected_file = class_var.get()
    if not selected_file:
        messagebox.showerror("Error", "Please select a file.")
        return

    present = [name for name, var in checkboxes.items() if var.get() == 1]
    result = mark_attendance(selected_file, present)

    if result == "Marked successfully":
        messagebox.showinfo("Success", "Attendance marked!")
    elif result == "Attendance already marked":
        messagebox.showwarning("Already Marked", "You already marked today.")
    else:
        messagebox.showerror("Error", result)

    for var in checkboxes.values():
        var.set(0)

# Merge all students into one file
def merge_students():
    output_file = merge_all_students()
    messagebox.showinfo("Merged", f"All students saved to {output_file}")

# Buttons
tk.Button(root, text="Add Student", command=add_name, bg="orange", fg="white", font=("Arial", 12)).pack(pady=5)
tk.Button(root, text="Load Students", command=load_students, bg="blue", fg="white", font=("Arial", 12)).pack(pady=5)
tk.Button(root, text="Mark Attendance", command=submit_attendance, bg="green", fg="white", font=("Arial", 12)).pack(pady=5)
tk.Button(root, text="Merge All Students", command=merge_students, bg="purple", fg="white", font=("Arial", 12)).pack(pady=15)

root.mainloop()
